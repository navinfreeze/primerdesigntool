import urllib3
import getpass
import json
import requests
import time
import openpyxl

# -------------------------------------------------------------------------------------------------------------
# Functions

# GC percent


def gcpercent(sequence):
    count = {'A': 0, 'T': 0, 'G': 0, 'C': 0}
    for base in sequence:
        count[base] += 1
    gc_content = ((count['C']+count['G'])/len(sequence))*100
    return (gc_content)

# Mismatch Alignment


def naive(p, t, m):  # p - primer, t- genome , m - maximum mismatches allowed
    occurences = 0
    for i in range(len(t) - len(p) + 1):
        match = True
        mis = 0
        for j in range(len(p)):
            if not t[j+i] == p[j]:
                mis += 1
                if mis > m:
                    match = False
                    break
        if match:
            occurences += 1
    return occurences  # returns the number of such alignments

# Reverse Complement_DNA


def revcomp(s1):
    a = 0
    s2 = ''
    complement = {'A': 'T', 'T': 'A', 'G': 'C', 'C': 'G'}
    while a < len(s1):
        s2 = complement[s1[a]] + s2
        a += 1
    return s2

# Complement_DNA


def comp(s1):
    a = 0
    s2 = ''
    complement = {'A': 'T', 'T': 'A', 'G': 'C', 'C': 'G'}
    while a < len(s1):
        s2 = s2 + complement[s1[a]]
        a += 1
    return s2
# -------------------------------------------------------------------------------------------------------------
# Taking Inputs


seq 			= 	input("Enter gene/DNA sequence: ")
print ("The index of the first nucleotide in the above sequence is taken as one")
print ("The length of the total sequence is" + str(len(seq)))
forward_primer_start 	= 	int(input("Enter index of forward primer start: "))
forward_primer_end 	= 	int(input("Enter index of forward primer end: "))
reverse_primer_start 	=  	int(input("Enter index of reverse primer end: "))
reverse_primer_end 	= 	int(input("Enter index of reverse primer end: "))
min_primer_length 	= 	int(input("Enter minimum primer Length (recommended: 18): "))
max_primer_length 	= 	int(input("Enter maximum primer Length (recommended: 20): "))
max_gcpercent 		= 	int(input("Enter maximum GC Percent (recommended: 60): "))
min_melt_temp 		= 	int(input("Enter minimum melting Temp (recommended: 55): "))
min_selfdimer 		= 	int(input("Enter maximum delta G for self-dimer (recommended: -9):"))
min_hairpin 		= 	int(input("Enter maximum delta G for hairpin (recommended: -9):"))
max_mismatch 		= 	int(input("Enter maximum mismatches (recommended: 6):"))
fiveprimeflank 	= 	input("Enter sequence fiveprime flank sequence (Leave empty if not required): ")
threeprimeflank 	= 	input("Enter sequence threeprime flank sequence (Leave empty if not required): ")


if type(forward_primer_start) == str:
    forward_primer_start = 0
if type(forward_primer_end) == str:
    forward_primer_end = len(seq)
if type(reverse_primer_start) == str:
    reverse_primer_start = 0
if type(reverse_primer_end) == str:
    reverse_primer_end = len(seq)
if type(min_primer_length) == str:
    min_primer_length = 18
if type(max_primer_length) == str:
    max_primer_length = 25
if type(max_gcpercent) == str:
    max_gcpercent = 65
if type(min_melt_temp) == str:
    min_melt_temp = 55
if type(min_selfdimer) == str:
    min_selfdimer = -9
if type(min_hairpin) == str:
    min_hairpin = -9
if type(max_mismatch) == str:
    max_mismatch = 10
if type(fiveprimeflank) == str:
    fiveprimeflank = ""
if type(threeprimeflank) == str:
    threeprimeflank = ""

print("")
print("Input parameters")
print("")
print("Forward primer from " + str(forward_primer_start) + " to " + str(forward_primer_end))
print("Forward primer from " + str(reverse_primer_start) + " to " + str(reverse_primer_end))
print("Primer length from " + str(min_primer_length) + " to " + str(max_primer_length))
print("Max GC percent = " + str(max_gcpercent))
print("Min melt temp = " + str(min_melt_temp))
print("Max selfdimer = " + str(min_selfdimer))
print("Max hairpin = " + str(min_hairpin))
print("Max mismatch = " + str(max_mismatch))
print("Fiveprimeflank = " + str(fiveprimeflank))
print("Threeprimeflank = " + str(threeprimeflank))

seq = seq.upper()
# Finding The DNA sequence from coding RNA strand
sequence = seq
seq = ""
for i in range(0, len(sequence)):
    if sequence[i] == "A":
        seq = seq+"A"
    elif sequence[i] == "U":
        seq = seq+"T"
    elif sequence[i] == "T":
        seq = seq+"T"
    elif sequence[i] == "G":
        seq = seq+"G"
    elif sequence[i] == "C":
        seq = seq+"C"
revseq = revcomp(seq)
seqlen = len(seq)
fiveprimeflank = fiveprimeflank.upper()
threeprimeflank = threeprimeflank.upper()

print("This may take a While Please Wait while it is Loading")

startlog = time.time()
# -------------------------------------------------------------------------------------------------------------
# Login to API


def api_login():
    urllib3.disable_warnings()  # To disable Insecurity warnings
    # Token Generation and Authorization for the IDT API OligoAnalyzer Tool
    token_url = "https://www.idtdna.com/IdentityServer/connect/token"
    RO_user = 'Navinchandra'
    RO_password = 'Navinpravin@4'
    client_id = 'Navinchandra'
    client_secret = '5ec8bcd1-ec3e-4f8b-af96-9cc26f0db3c3'
    data = {'grant_type': 'password',
            'username': 'Navinchandra', 'password': 'Navinpravin@4', 'scope': 'test'}
    access_token_response = requests.post(
        token_url, data=data, verify=False, allow_redirects=False, auth=(client_id, client_secret))
    tokens = json.loads(access_token_response.text)
    headers = {
        'Content-Type': 'application/json',
        'Authorization': 'Bearer ' + tokens['access_token'],
        'Cookie': 'ARRWestffinity=851675575f498916558244b4dad1bb32e577e75c4f3a4a0c957416617471c181'
    }
    return (headers)


headers = api_login()
# -------------------------------------------------------------------------------------------------------------
# Functions for primer analysis


def melt_temp(primer):
    url = "https://www.idtdna.com/restapi/v1/OligoAnalyzer/Analyze"
    payload = str('{\"Sequence\":\"') + primer + str(
        '\",\"NaConc\":50,\"MgConc\":0,\"DNTPsConc\":0,\"OligoConc\":0.25,\"NucleotideType\":\"DNA\"}')
    response = requests.request("POST", url, headers=headers, data=payload)
    melt_temp = response.json()['MeltTemp']
    return (melt_temp)


def selfdimer(primer):
    url = "https://www.idtdna.com/restapi/v1/OligoAnalyzer/HeteroDimer?primary=" + \
        primer + "&secondary=" + primer
    payload = {}
    response = requests.request("POST", url, headers=headers, data=payload)
    selfdimer = response.json()[0]['DeltaG']
    return (selfdimer)


def hairpin(primer):
    url = "https://www.idtdna.com/restapi/v1/OligoAnalyzer/Hairpin"
    payload = str('{\r\n  \"Sequence\": \"') + primer + str(
        '\",\r\n  \"NaConc\": 50,\r\n  \"FoldingTemp\": 25,\r\n  \"MgConc\": 0,\r\n  \"NucleotideType\": \"DNA\"\r\n}')
    response = requests.request("POST", url, headers=headers, data=payload)
    hairpin = response.json()[0]['deltaG']
    return (hairpin)


# -------------------------------------------------------------------------------------------------------------
# IDT Oligo Analyzer Automation - GC, Melt Temp, Homodimer for both Forward And Reverse Primers
# -------------------------------------------------------------------------------------------------------------
# Open Excel containing Primers
wb = openpyxl.Workbook()
# Create New Sheet Containing all Values
wb.create_sheet('Shortlist Primers')
sheet = wb['Shortlist Primers']
sheet.cell(row=1, column=1).value = 'Forward Primer Name'
sheet.cell(row=1, column=2).value = 'Forward Primer Sequence'
sheet.cell(row=1, column=3).value = 'Forward Primer GC Content'
sheet.cell(row=1, column=4).value = 'Forward Primer Melt Temp'
sheet.cell(row=1, column=5).value = 'Forward Primer Hairpin Delta G'
sheet.cell(row=1, column=6).value = 'Forward Primer Self-Dimer Delta G'
sheet.cell(row=1, column=7).value = 'Forward Primer Without Flank Sequence'
sheet.cell(row=1, column=8).value = 'Forward Primer Without Flank GC Content'
sheet.cell(row=1, column=9).value = 'Forward Primer Without Flank Melt Temp'

sheet.cell(row=1, column=11).value = 'Reverse Primer Name'
sheet.cell(row=1, column=12).value = 'Reverse Primer Sequence'
sheet.cell(row=1, column=13).value = 'Reverse Primer GC Content'
sheet.cell(row=1, column=14).value = 'Reverse Primer Melt Temp'
sheet.cell(row=1, column=15).value = 'Reverse Primer Hairpin Delta G'
sheet.cell(row=1, column=16).value = 'Reverse Primer Self-Dimer Delta G'
sheet.cell(row=1, column=17).value = 'Reverse Primer Without Flank Sequence'
sheet.cell(row=1, column=18).value = 'Reverse Primer Without Flank GC Content'
sheet.cell(row=1, column=19).value = 'Reverse Primer Without Flank Melt Temp'
# -------------------------------------------------------------------------------------------------------------
# Forward Primers
print("Analyzing Forward Primers")
count = 1
for k in range(0, forward_primer_end-forward_primer_start+1):
    for i in range(min_primer_length, max_primer_length+1):
        if time.time() - 2000 > startlog:
            headers = api_login()
            startlog = time.time()
            print("Re-login Successul")
        primer = fiveprimeflank + \
            seq[forward_primer_start+k:forward_primer_start+k+i]
        primer_without_flank = seq[forward_primer_start +
                                   k:forward_primer_start+k+i]
        if primer[-1] == 'G' or primer[-1] == 'C':
            gc1 = gcpercent(primer)
            gc2 = gcpercent(primer_without_flank)
            if gc1 < max_gcpercent and gc2 < max_gcpercent:
                if naive(primer, seq, max_mismatch) <= 1 and naive(primer, revseq, max_mismatch) <= 1:
                    tm1 = melt_temp(primer)
                    if fiveprimeflank == "":
                        tm2 = tm1
                    else:
                        tm2 = melt_temp(primer_without_flank)
                    if tm1 > min_melt_temp and tm2 > min_melt_temp:
                        sd = selfdimer(primer)
                        if sd > min_selfdimer:
                            hp = hairpin(primer)
                            if hp > min_hairpin:
                                primer_name = 'forprimer' + '_' + \
                                    str(forward_primer_start+k+1) + \
                                    '_' + str(len(fiveprimeflank)+i)
                                count += 1
                                sheet.cell(
                                    row=count, column=1).value = primer_name
                                sheet.cell(row=count, column=2).value = primer
                                sheet.cell(row=count, column=3).value = gc1
                                sheet.cell(row=count, column=4).value = tm1
                                sheet.cell(row=count, column=5).value = hp
                                sheet.cell(row=count, column=6).value = sd
                                sheet.cell(
                                    row=count, column=7).value = primer_without_flank
                                sheet.cell(row=count, column=8).value = gc2
                                sheet.cell(row=count, column=9).value = tm2
# -------------------------------------------------------------------------------------------------------------
# Reverse Primers
print("Analyzing Reverse Primers")
count = 1
for k in range(0, reverse_primer_end-reverse_primer_start+1):
    for i in range(min_primer_length, max_primer_length+1):
        if time.time() - 2000 > startlog:
            startlog = time.time()
            headers = api_login()
            print("Re-login Successul")
        primer = revcomp(threeprimeflank) + \
            revseq[seqlen-reverse_primer_end+k:seqlen-reverse_primer_end+k+i]
        primer_without_flank = revseq[seqlen -
                                      reverse_primer_end+k:seqlen-reverse_primer_end+k+i]
        if primer[-1] == 'G' or primer[-1] == 'C':
            gc1 = gcpercent(primer)
            gc2 = gcpercent(primer_without_flank)
            if gc1 < max_gcpercent and gc2 < max_gcpercent:
                if naive(primer, seq, max_mismatch) <= 1 and naive(primer, revseq, max_mismatch) <= 1:
                    tm1 = melt_temp(primer)
                    if threeprimeflank == "":
                        tm2 = tm1
                    else:
                        tm2 = melt_temp(primer_without_flank)
                    if tm1 > min_melt_temp and tm2 > min_melt_temp:
                        sd = selfdimer(primer)
                        if sd > min_selfdimer:
                            hp = hairpin(primer)
                            if hp > min_hairpin:
                                primer_name = 'revprimer' + '_' + \
                                    str(reverse_primer_end-k-i+1) + \
                                    '_' + str(len(threeprimeflank)+i)
                                count += 1
                                sheet.cell(
                                    row=count, column=11).value = primer_name
                                sheet.cell(row=count, column=12).value = primer
                                sheet.cell(row=count, column=13).value = gc1
                                sheet.cell(row=count, column=14).value = tm1
                                sheet.cell(row=count, column=15).value = hp
                                sheet.cell(row=count, column=16).value = sd
                                sheet.cell(
                                    row=count, column=17).value = primer_without_flank
                                sheet.cell(row=count, column=18).value = gc2
                                sheet.cell(row=count, column=19).value = tm2
wb.save('primers_analyzed.xlsx')
# -------------------------------------------------------------------------------------------------------------
print("Open the Excel File to view the Primers")
